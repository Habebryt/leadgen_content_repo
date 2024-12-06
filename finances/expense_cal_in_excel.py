import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

def create_transaction_sheet(filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Transactions"

    headers = ["Date", "Transaction Information", "Credit", "Debit", "Account Balance"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Set column widths
    column_widths = [15, 40, 15, 15, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Add a starting balance row
    sheet.cell(row=2, column=1, value=datetime.now().strftime("%Y-%m-%d"))
    sheet.cell(row=2, column=2, value="Starting Balance")
    sheet.cell(row=2, column=5, value=0)

    # Set up formulas for automatic balance calculation
    sheet.cell(row=3, column=5, value="=E2+C3-D3")
    sheet.cell(row=3, column=5).number_format = '#,##0.00'

    # Create a table
    tab = Table(displayName="TransactionsTable", ref="A1:E3")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    wb.save(filename)
    print(f"Created new Excel file: {filename}")

def add_sample_transactions(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    transactions = [
        ("2023-10-01", "Initial Deposit", 10000, 0),
        ("2023-10-05", "Grocery Shopping", 0, 150),
        ("2023-10-10", "Salary", 5000, 0),
        ("2023-10-15", "Utility Bill", 0, 200),
    ]

    for transaction in transactions:
        next_row = sheet.max_row + 1
        for col, value in enumerate(transaction, start=1):
            sheet.cell(row=next_row, column=col, value=value)
        
        # Copy down the balance formula
        balance_cell = sheet.cell(row=next_row, column=5)
        balance_cell.value = f"=E{next_row-1}+C{next_row}-D{next_row}"
        balance_cell.number_format = '#,##0.00'

    # Extend the table to include new rows
    tab = sheet.tables["TransactionsTable"]
    tab.ref = f"A1:E{sheet.max_row}"

    wb.save(filename)
    print("Added sample transactions")

def main():
    filename = "financial_transactions.xlsx"
    create_transaction_sheet(filename)
    add_sample_transactions(filename)
    print(f"Exported workbook to {filename}")
    print("You can now open this file in Microsoft Excel and add more transactions directly.")

if __name__ == "__main__":
    main()