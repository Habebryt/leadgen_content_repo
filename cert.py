import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

def create_invoice_excel(filename="invoice.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Header
    ws['A1'] = "Customer"
    ws['A2'] = "Customer email"
    ws['A3'] = "Cc(3)/Bcc"
    ws['A4'] = "Send later"
    ws['A5'] = "BALANCE DUE"
    ws['B5'] = "₦525,000.00"
    ws['A6'] = "For your information"
    ws['B6'] = "You can't add data because your QuickBooks subscription isn't active. Please update your subscription or contact your administrator."
    ws['A7'] = "Billing address"
    ws['A8'] = "Terms"
    ws['A9'] = "Invoice date"
    ws['A10'] = "Due date"
    ws['A11'] = "Invoice no."
    ws['A12'] = "Tags"

    # Apply styles to header
    for row in range(1, 13):
        ws[f'A{row}'].font = Font(bold=True)

    ws['B5'].font = Font(bold=True)

    # Invoice items
    headers = ["#", "Service Date", "Product/Service", "Description", "Qty", "Rate", "Amount"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=14, column=col, value=header).font = Font(bold=True)

    items = [
        (1, "10/08/2024", "BIDs:ITF", "Industrial Training Fund (ITF) Compliance Certificate", 1, 375000, 375000),
        (2, "10/08/2024", "NSITF", "National Social Insurance Trust Fund (NSITF) Compliance Certificate", 1, 80000, 80000),
        (3, "10/08/2024", "Taxes:TCC", "Tax Clearance Certificate", 1, 70000, 70000)
    ]

    for row, item in enumerate(items, start=15):
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)

    # Totals
    ws['F18'] = "Total"
    ws['G18'] = "=SUM(G15:G17)"

    # Apply border to the table
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws['A14:G18']:
        for cell in row:
            cell.border = border

    # Save the workbook
    wb.save(filename)
    print(f"Invoice saved as {filename}")

# Run the function
create_invoice_excel()