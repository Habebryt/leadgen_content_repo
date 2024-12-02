import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

def create_craffta_training_budget(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Craffta Training Budget"

    def style_cell(cell, border=Border(), fill=None, font=None, alignment=None):
        cell.border = border
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment

    def create_header(sheet, row, headers):
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            style_cell(cell, 
                       border=Border(bottom=Side(style='medium')),
                       font=Font(bold=True, color="FFFFFF"),
                       fill=PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                       alignment=Alignment(horizontal="center", vertical="center"))

    # Budget data
    budget_data = [
        ("Recruitment Costs", "", "", ""),
        ("Job Postings", 12, 500, "=B3*C3"),
        ("Recruitment Agency Fees", 4, 2000, "=B4*C4"),
        ("Background Checks", 12, 100, "=B5*C5"),
        ("Salary Scales", "", "", ""),
        ("Junior Trainer (2)", 12, 3000, "=B7*C7*2"),
        ("Senior Trainer (1)", 12, 5000, "=B8*C8"),
        ("Training Manager (1)", 12, 7000, "=B9*C9"),
        ("Allowances", "", "", ""),
        ("Housing Allowance (10% of base)", 12, "=0.1*(C7*2+C8+C9)", "=B11*C11"),
        ("Transportation Allowance", 12, 200, "=B12*C12*4"),
        ("Operational Costs", "", "", ""),
        ("Office Rent", 12, 2000, "=B14*C14"),
        ("Utilities (5% of base)", 12, "=0.05*(C7*2+C8+C9)", "=B15*C15"),
        ("Office Supplies", 12, 300, "=B16*C16"),
        ("Travel Expenses", "", "", ""),
        ("Domestic Travel", 12, 1000, "=B18*C18"),
        ("International Travel", 4, 3000, "=B19*C19"),
        ("Data and Communication", "", "", ""),
        ("Internet", 12, 200, "=B21*C21"),
        ("Mobile Data Plans", 12, 100, "=B22*C22*4"),
        ("Technology Budget", "", "", ""),
        ("Hardware (Laptops, Projectors)", 1, 10000, "=B24*C24"),
        ("Software Licenses", 12, 500, "=B25*C25"),
        ("Training Materials", "", "", ""),
        ("Printed Materials", 12, 500, "=B27*C27"),
        ("Online Resources", 12, 300, "=B28*C28"),
        ("Marketing and Advertising", "", "", ""),
        ("Digital Marketing", 12, 1000, "=B30*C30"),
        ("Print Advertising", 4, 2000, "=B31*C31"),
        ("Payroll Taxes (15% of total salary)", 12, "=0.15*(D7+D8+D9)", "=B33*C33"),
        ("Employee Benefits (10% of total salary)", 12, "=0.1*(D7+D8+D9)", "=B34*C34"),
        ("Miscellaneous (5% of total budget)", 1, "=0.05*D37", "=B35*C35")
    ]

    # Create headers
    headers = ["Expense Category", "Frequency (Monthly)", "Monthly Amount", "Annual Amount"]
    create_header(ws, 1, headers)

    # Populate data
    for row, (category, frequency, amount, formula) in enumerate(budget_data, start=2):
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=frequency)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=4, value=formula)

        # Style data cells
        for col in range(1, 5):
            style_cell(ws.cell(row=row, column=col),
                       border=Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                     left=Side(style='thin'), right=Side(style='thin')))
        
        # Bold category headers
        if frequency == "":
            ws.cell(row=row, column=1).font = Font(bold=True)

    # Add total row
    total_row = len(budget_data) + 2
    ws.cell(row=total_row, column=1, value="Total Annual Budget")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")

    # Style total row
    for col in range(1, 5):
        style_cell(ws.cell(row=total_row, column=col),
                   border=Border(top=Side(style='double'), bottom=Side(style='double'), 
                                 left=Side(style='thin'), right=Side(style='thin')),
                   font=Font(bold=True))

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Format numbers
    for row in range(2, total_row + 1):
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'].number_format = '$#,##0'

    # Add conditional formatting
    ws.conditional_formatting.add(f'D2:D{total_row-1}',
                                  CellIsRule(operator='greaterThan',
                                             formula=['=$D$37*0.05'],
                                             stopIfTrue=True,
                                             fill=PatternFill(start_color='FFC7CE',
                                                              end_color='FFC7CE',
                                                              fill_type='solid')))

    # Create a bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Annual Expenses by Category"
    chart.y_axis.title = 'Amount'
    chart.x_axis.title = 'Category'

    data = Reference(ws, min_col=4, min_row=1, max_row=total_row-1, max_col=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=total_row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "F2")

    # Save the workbook
    wb.save(filename)

# Create the Excel file
create_craffta_training_budget("Craffta_Training_Budget.xlsx")

print("Excel spreadsheet 'Craffta_Training_Budget.xlsx' has been created successfully.")